#!/usr/bin/env python3
"""
Pilot automation runner.

Reads the exported competition spreadsheet, matches rows against a list of
automation targets, and invokes the shared Selenium workflow for each match.

This is intentionally conservative: you must opt-in by providing a
`automation_targets.json` file with explicit link substrings that are safe to
automate. Use `--dry-run` while validating.
"""

from __future__ import annotations

import argparse
import json
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook

from autofill_core import FillAction, load_json, perform_autofill
from state_utils import load_state, save_state

logger = logging.getLogger("auto_entry_runner")


@dataclass
class AutomationTarget:
    match: str
    config_path: Path
    data_path: Path
    screenshot_dir: Optional[Path] = None
    submit_selector: Optional[str] = None

    def matches(self, link: str) -> bool:
        return self.match in link


def load_targets(path: Path) -> List[AutomationTarget]:
    if not path.exists():
        raise FileNotFoundError(f"Automation target file not found: {path}")
    raw = json.loads(path.read_text(encoding="utf-8"))
    defaults = raw.get("defaults", {})
    targets_data = raw.get("targets", [])
    targets: List[AutomationTarget] = []
    for entry in targets_data:
        match = entry.get("match")
        config = entry.get("config", defaults.get("config"))
        data = entry.get("data", defaults.get("data"))
        if not (match and config and data):
            logger.warning("Skipping malformed target entry: %s", entry)
            continue
        screenshot_dir = entry.get("screenshot_dir", defaults.get("screenshot_dir"))
        submit_selector = entry.get("submit_selector", defaults.get("submit_selector"))
        targets.append(
            AutomationTarget(
                match=match,
                config_path=Path(config),
                data_path=Path(data),
                screenshot_dir=Path(screenshot_dir) if screenshot_dir else None,
                submit_selector=submit_selector,
            )
        )
    return targets


def auto_confirm(actions: List[FillAction], screenshot_path: Path, *, always_yes: bool) -> bool:
    print("\n=== Autofill preview ===")
    for action in actions:
        label = (action.label or "<no label>").strip()
        print(
            f"- label='{label[:60]}' type={action.input_type} key={action.mapped_key} "
            f"filled={action.filled} value={action.value}"
        )
    print(f"Screenshot -> {screenshot_path}")
    if always_yes:
        print("Auto-confirm enabled; submitting.")
        return True
    response = input("Submit this form automatically? [yes/No]: ").strip().lower()
    return response == "yes"


def run(
    workbook_path: Path,
    targets_path: Path,
    *,
    dry_run: bool,
    auto_yes: bool,
    limit: Optional[int],
) -> List[str]:
    targets = load_targets(targets_path)
    if not targets:
        logger.warning("No automation targets defined; nothing to do.")
        return []

    wb = load_workbook(workbook_path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        link_idx = headers.index("Link")
        status_idx = headers.index("Successful Submission")
    except ValueError as exc:
        raise ValueError("Spreadsheet missing expected columns (Link / Successful Submission).") from exc
    try:
        new_idx = headers.index("Is New This Run")
    except ValueError:
        new_idx = -1

    processed = 0
    successes = 0
    submitted_links: List[str] = []
    for row in ws.iter_rows(min_row=2):
        link = row[link_idx].value or ""
        if not link:
            continue
        target = next((t for t in targets if t.matches(link)), None)
        if not target:
            continue

        if limit is not None and processed >= limit:
            break

        processed += 1
        logger.info("Processing competition %s", link)

        if dry_run:
            print(f"[DRY RUN] Would submit: {link} using {target.config_path}")
            continue

        cfg = load_json(target.config_path)
        data = load_json(target.data_path)
        cfg["url"] = link
        if target.screenshot_dir:
            target.screenshot_dir.mkdir(parents=True, exist_ok=True)
            cfg["screenshot_dir"] = str(target.screenshot_dir)
        if target.submit_selector:
            cfg["submit_selector"] = target.submit_selector

        outcome = perform_autofill(
            cfg,
            data,
            confirm_submit=lambda actions, screenshot: auto_confirm(actions, screenshot, always_yes=auto_yes),
        )

        if outcome.submitted:
            successes += 1
            submitted_links.append(link)
            row[status_idx].value = "Yes"
            if new_idx >= 0:
                row[new_idx].value = (row[new_idx].value or "").replace("YES", "").strip()
        else:
            row[status_idx].value = row[status_idx].value or ""

    if not dry_run:
        wb.save(workbook_path)
        logger.info("Updated spreadsheet with %d successful submissions.", successes)
    logger.info("Processed %d competitions (%d successes).", processed, successes)
    return submitted_links


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Automate competition entries for whitelisted targets.")
    parser.add_argument(
        "-e",
        "--entries",
        type=Path,
        default=Path("competition_entries.xlsx"),
        help="Path to the competition spreadsheet.",
    )
    parser.add_argument(
        "-t",
        "--targets",
        type=Path,
        default=Path("automation_targets.json"),
        help="Automation target definition file.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="List matching competitions without launching Selenium.",
    )
    parser.add_argument(
        "--auto-confirm",
        action="store_true",
        help="Skip the manual confirmation prompt and submit automatically.",
    )
    parser.add_argument(
        "-n",
        "--limit",
        type=int,
        help="Maximum number of competitions to process this run.",
    )
    parser.add_argument(
        "--state",
        type=Path,
        default=Path("competition_state.json"),
        help="State file to mark successful submissions (matches discovery script).",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="Enable verbose logging.",
    )
    return parser


def main(argv: Optional[List[str]] = None) -> None:
    parser = build_parser()
    args = parser.parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s: %(message)s",
    )
    submitted = run(
        args.entries,
        args.targets,
        dry_run=args.dry_run,
        auto_yes=args.auto_confirm,
        limit=args.limit,
    )
    if submitted:
        logger.info("Successfully submitted %d competitions.", len(submitted))
    if submitted and not args.dry_run and args.state:
        state = load_state(args.state)
        state.submitted.update(submitted)
        state.seen.update(submitted)
        save_state(args.state, state)


if __name__ == "__main__":
    main()
